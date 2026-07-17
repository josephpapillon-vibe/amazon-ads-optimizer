#!/usr/bin/env python3
"""Rebuild the reporting database from client folders and export a dashboard snapshot.

Run from the project root: python3 db/build_db.py

Sources (read-only — this script never writes back to clients/):
  - clients/<client>/config.json           -> target_acos per client
  - clients/<client>/logs/changes_*.csv     -> every past optimize.py batch/decision
  - clients/<client>/context/*.xlsx         -> monthly per-product sales + ROI targets
    (only jmn has this report today; other clients are skipped if absent)

Outputs:
  - db/optimizer.db   SQLite database, fully rebuilt each run (see schema.sql)
  - db/export.json    flat snapshot of the db, consumed by the dashboard artifact

Note: account_aov / baseline_cvr are computed by optimize.py at run time but not
currently written to logs/changes_*.csv, so historical batches in this db have
those fields as NULL. Only a future batch would carry them, if optimize.py is
changed to persist them.
"""
import csv
import glob
import json
import os
import re
import sqlite3
import unicodedata

import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CLIENTS_DIR = os.path.join(BASE, "clients")
DB_PATH = os.path.join(BASE, "db", "optimizer.db")
SCHEMA_PATH = os.path.join(BASE, "db", "schema.sql")
EXPORT_PATH = os.path.join(BASE, "db", "export.json")

FR_MONTHS = {
    "janvier": 1, "fevrier": 2, "mars": 3, "avril": 4, "mai": 5, "juin": 6,
    "juillet": 7, "aout": 8, "septembre": 9, "octobre": 10, "novembre": 11,
    "decembre": 12,
}


def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def month_num(name):
    return FR_MONTHS.get(strip_accents(str(name)).strip().lower())


def normalize_market(label):
    label = strip_accents(str(label)).strip().lower()
    if label.startswith("canada"):
        return "CAD"
    if label.startswith("etats-unis") or label.startswith("etats unis"):
        return "USD"
    return None


def parse_tier(reason):
    if not reason:
        return None
    if "extreme" in reason:
        return "extreme"
    if "normal" in reason:
        return "normal"
    return None


def ingest_bid_decisions(conn, client_id):
    log_dir = os.path.join(CLIENTS_DIR, client_id, "logs")
    for path in sorted(glob.glob(os.path.join(log_dir, "changes_*.csv"))):
        date_part = os.path.basename(path)[len("changes_"):-len(".csv")]
        cur = conn.execute(
            "INSERT INTO batches (client_id, batch_date, account_aov, baseline_cvr) VALUES (?, ?, NULL, NULL)",
            (client_id, date_part),
        )
        batch_id = cur.lastrowid
        with open(path, newline="") as f:
            for row in csv.DictReader(f):
                old_bid = float(row["old_bid"]) if row.get("old_bid") else None
                new_bid = float(row["new_bid"]) if row.get("new_bid") else None
                direction = "hold"
                if old_bid is not None and new_bid is not None:
                    direction = "up" if new_bid > old_bid else ("down" if new_bid < old_bid else "hold")
                conn.execute(
                    """INSERT INTO bid_decisions
                       (batch_id, action, entity, target_id, campaign, ad_group, target_text,
                        old_bid, new_bid, clicks, spend, sales, orders, reason, tier, direction)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        batch_id, row.get("action"), row.get("entity"), row.get("id"),
                        row.get("campaign"), row.get("ad_group"), row.get("target"),
                        old_bid, new_bid,
                        float(row["clicks"]) if row.get("clicks") else None,
                        float(row["spend"]) if row.get("spend") else None,
                        float(row["sales"]) if row.get("sales") else None,
                        float(row["orders"]) if row.get("orders") else None,
                        row.get("reason"), parse_tier(row.get("reason")), direction,
                    ),
                )


def find_report_workbook(client_id):
    context_dir = os.path.join(CLIENTS_DIR, client_id, "context")
    for path in glob.glob(os.path.join(context_dir, "*.xlsx")):
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception:
            continue
        if "ROI target" in wb.sheetnames or any("Ventes par produit" in (s or "") for s in wb.sheetnames):
            return path
    return None


def parse_product_sales(wb):
    """Returns {(market, sku, year, month): sales}. Later sheets (by as-of date)
    overwrite earlier ones since each sheet reports a rolling 6-month window and
    later reports are the more corrected/final figures for a given month."""
    sheet_re = re.compile(r"^(\d{4})\s*-\s*(.+)$")
    dated_sheets = []
    for name in wb.sheetnames:
        m = sheet_re.match(name.strip())
        if not m:
            continue
        year, month_name = int(m.group(1)), month_num(m.group(2))
        if month_name is None:
            continue
        dated_sheets.append((year, month_name, name))
    dated_sheets.sort()

    result = {}
    for year, month, name in dated_sheets:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = next((i for i, r in enumerate(rows) if r[1] and "Ventes par produit" in str(r[1])), None)
        if header_idx is None:
            continue
        month_row = rows[header_idx + 1]
        months = []
        col = 3
        while col < len(month_row) and month_row[col]:
            mn = month_num(month_row[col])
            if mn is None:
                break
            months.append(mn)
            col += 1
        if not months:
            continue
        # months list ends at this sheet's as-of month; walk backward to assign (year, month) per column
        col_dates = []
        y, mo = year, month
        for _ in range(len(months)):
            col_dates.insert(0, (y, mo))
            mo -= 1
            if mo == 0:
                mo, y = 12, y - 1

        current_market = None
        for r in rows[header_idx + 2:]:
            label = r[1] if len(r) > 1 else None
            if label and "Resume" in strip_accents(str(label)):
                break
            if label:
                mkt = normalize_market(label)
                if mkt:
                    current_market = mkt
            sku = r[2] if len(r) > 2 else None
            if not sku or current_market is None:
                continue
            for i, (y2, mo2) in enumerate(col_dates):
                val = r[3 + i] if 3 + i < len(r) else None
                if not isinstance(val, (int, float)):
                    continue
                result[(current_market, sku, y2, mo2)] = float(val)
    return result


def parse_roi_targets(wb):
    if "ROI target" not in wb.sheetnames:
        return {}
    ws = wb["ROI target"]
    rows = list(ws.iter_rows(values_only=True))
    result = {}
    for r in rows[2:]:
        sku = r[0] if r else None
        if not sku:
            break
        for bucket, val in (("Q4", r[1] if len(r) > 1 else None), ("Q1-2-3", r[2] if len(r) > 2 else None)):
            if isinstance(val, (int, float)):
                result[(sku, bucket)] = float(val)
    return result


def main():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    with open(SCHEMA_PATH) as f:
        conn.executescript(f.read())

    clients = sorted(
        d for d in os.listdir(CLIENTS_DIR) if os.path.isdir(os.path.join(CLIENTS_DIR, d))
    )
    for client_id in clients:
        cfg_path = os.path.join(CLIENTS_DIR, client_id, "config.json")
        if not os.path.exists(cfg_path):
            continue
        with open(cfg_path) as f:
            cfg = json.load(f)
        conn.execute("INSERT INTO clients (client_id, target_acos) VALUES (?, ?)",
                     (client_id, cfg.get("target_acos")))
        ingest_bid_decisions(conn, client_id)

        wb_path = find_report_workbook(client_id)
        if wb_path:
            wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
            sales = parse_product_sales(wb)
            for (market, sku, year, month), value in sales.items():
                conn.execute(
                    "INSERT OR REPLACE INTO product_sales (client_id, sku, market, year, month, sales) VALUES (?,?,?,?,?,?)",
                    (client_id, sku, market, year, month, value),
                )
            roi = parse_roi_targets(wb)
            for (sku, bucket), value in roi.items():
                conn.execute(
                    "INSERT OR REPLACE INTO product_roi_targets (client_id, sku, bucket, roi_target) VALUES (?,?,?,?)",
                    (client_id, sku, bucket, value),
                )
            print(f"{client_id}: ingested {len(sales)} product-sales cells, {len(roi)} ROI targets from {os.path.basename(wb_path)}")
        conn.commit()
        print(f"{client_id}: batches + decisions ingested")

    export_snapshot(conn)
    conn.close()
    print(f"\nDatabase rebuilt at {DB_PATH}")
    print(f"Dashboard snapshot exported to {EXPORT_PATH}")


def export_snapshot(conn):
    conn.row_factory = sqlite3.Row
    data = {"clients": {}}
    for client in conn.execute("SELECT * FROM clients"):
        cid = client["client_id"]
        batches = [dict(b) for b in conn.execute(
            "SELECT * FROM batches WHERE client_id = ? ORDER BY batch_date", (cid,)
        )]
        for b in batches:
            b["decisions"] = [dict(d) for d in conn.execute(
                "SELECT * FROM bid_decisions WHERE batch_id = ?", (b["batch_id"],)
            )]
        sales = [dict(s) for s in conn.execute(
            "SELECT sku, market, year, month, sales FROM product_sales WHERE client_id = ? ORDER BY sku, market, year, month",
            (cid,),
        )]
        roi = [dict(r) for r in conn.execute(
            "SELECT sku, bucket, roi_target FROM product_roi_targets WHERE client_id = ?", (cid,)
        )]
        data["clients"][cid] = {
            "target_acos": client["target_acos"],
            "batches": batches,
            "product_sales": sales,
            "roi_targets": roi,
        }
    with open(EXPORT_PATH, "w") as f:
        json.dump(data, f, indent=1, default=str)


if __name__ == "__main__":
    main()
