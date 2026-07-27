#!/usr/bin/env python3
"""Amazon Ads bid optimizer.

Two-step workflow so nothing reaches the bulk upload file without a human checkpoint:
  1. python3 optimize.py <client-folder-name>
     Computes proposed bid changes and writes a review file — no bulk output yet.
  2. Review it (tools/review_app.py — double-click "Réviser un lot.command" — or hand-edit
     the review CSV's `approve` column), then:
     python3 optimize.py <client-folder-name> --apply <YYYY-MM-DD>
     Applies only the approved rows and writes the final bulk file + log.
"""
import sys
import json
import glob
import datetime
import csv
import os
import openpyxl

BID_ENTITIES = {"Keyword", "Product targeting"}
LOG_FIELDS = ["action", "entity", "id", "campaign", "ad_group", "target",
              "old_bid", "new_bid", "baseline_bid", "escalated",
              "clicks", "spend", "sales", "orders", "reason"]


def load_config(client_dir):
    with open(os.path.join(client_dir, "config.json")) as f:
        cfg = json.load(f)
    if cfg.get("target_acos") is None:
        sys.exit(f"target_acos is not set in {client_dir}/config.json — aborting.")
    return cfg


def find_input_file(client_dir):
    candidates = glob.glob(os.path.join(client_dir, "input", "*.xlsx"))
    if not candidates:
        sys.exit(f"No .xlsx file found in {client_dir}/input")
    if len(candidates) > 1:
        newest = max(candidates, key=os.path.getmtime)
        print(f"Multiple files in {client_dir}/input — using most recently modified: {os.path.basename(newest)}")
        return newest
    return candidates[0]


def find_campaigns_sheet(wb):
    for name in wb.sheetnames:
        if "Sponsored Products" in name and "Campaign" in name:
            return name
    sys.exit(f"Could not find a 'Sponsored Products Campaigns' sheet. Sheets present: {wb.sheetnames}")


def decide_bid(row, cfg, account_aov, baseline_cvr):
    """Returns (new_bid, reason, tier) where tier is 'normal', 'extreme', or None for no-change."""
    target_acos = cfg["target_acos"] / 100
    target_cpa = target_acos * account_aov if account_aov else None

    clicks = row["Clicks"] or 0
    spend = row["Spend"] or 0
    sales = row["Sales"] or 0
    orders = row["Orders"] or 0
    bid = row["Bid"]

    gate = cfg["data_gate"]
    if clicks < gate["min_clicks"]:
        return bid, "Insufficient data (below click threshold): no change", None

    if orders > 0:
        actual_acos = spend / sales if sales else 0
        low_cfg = cfg["bid_rules"]["low_acos"]
        high_cfg = cfg["bid_rules"]["high_acos"]
        low_threshold = target_acos * (low_cfg["threshold_pct_of_target"] / 100)
        extreme_low_threshold = target_acos * (low_cfg["extreme_pct_of_target"] / 100)

        if actual_acos > target_acos:
            extreme = actual_acos >= target_acos * high_cfg["extreme_multiple_of_target"]
            cap = (high_cfg["extreme_max_decrease_pct"] if extreme else high_cfg["normal_max_decrease_pct"]) / 100
            new_bid = bid * (target_acos / actual_acos)
            new_bid = max(new_bid, bid * (1 - cap))
            tier = "extreme" if extreme else "normal"
            return round(new_bid, 2), f"High ACOS ({actual_acos:.0%} > target {target_acos:.0%}, {tier}): bid trimmed up to {cap:.0%}", tier

        if actual_acos <= low_threshold:
            if actual_acos <= extreme_low_threshold:
                pct = low_cfg["extreme_max_increase_pct"]
                tier = "extreme"
            else:
                ratio = 1 - (actual_acos / low_threshold) if low_threshold else 1
                ratio = min(max(ratio, 0), 1)
                pct = low_cfg["normal_min_increase_pct"] + ratio * (low_cfg["normal_max_increase_pct"] - low_cfg["normal_min_increase_pct"])
                tier = "normal"
            new_bid = round(bid * (1 + pct / 100), 2)
            return new_bid, f"Low ACOS ({actual_acos:.0%} <= {low_threshold:.0%} threshold, {tier}): bid +{pct:.0f}%", tier

        return bid, f"Mid-band ACOS ({actual_acos:.0%}): no change", None

    # zero orders: require statistical confidence this isn't just normal variance,
    # not just a spend multiple. p_zero = chance an average keyword shows 0 orders
    # by chance alone at this click count, given the account's real conversion rate.
    wc = cfg["bid_rules"]["wasted_spend"]
    if baseline_cvr is None or baseline_cvr <= 0:
        return bid, "0 orders: no account-wide conversion rate available to assess confidence, no change", None

    p_zero = (1 - baseline_cvr) ** clicks
    if p_zero > wc["max_zero_order_probability_to_act"]:
        return bid, f"0 orders over {clicks:.0f} clicks: {p_zero:.0%} chance this is normal variance (baseline CVR {baseline_cvr:.1%}) — not enough confidence to cut yet", None

    extreme = p_zero <= wc["extreme_zero_order_probability"] and target_cpa is not None and spend >= wc["extreme_min_spend_multiple_of_target_cpa"] * target_cpa
    cap = (wc["extreme_max_decrease_pct"] if extreme else wc["normal_max_decrease_pct"]) / 100
    floor = wc["bid_floor"]
    new_bid = max(round(bid * (1 - cap), 2), floor)
    tier = "extreme" if extreme else "normal"
    return new_bid, f"0 orders over {clicks:.0f} clicks, spend ${spend:.2f} ({tier}, {p_zero:.0%} chance of pure variance): bid cut up to {cap:.0%}", tier


def roas_baseline_bid(bid, spend, sales, cfg):
    """Manual ROAS-based rule of thumb (validated by Joseph's colleague, 2026-07-20) used as the
    baseline bid whenever decide_bid finds a 'normal' tier change justified. Returns
    (new_bid, pct, label)."""
    rb = cfg["roas_baseline"]
    roas = sales / spend if spend else 0

    if spend < rb["min_spend_for_rules"]:
        pct = rb["low_spend_increase_pct"] / 100
        label = f"spend ${spend:.2f} < ${rb['min_spend_for_rules']}"
    elif roas > rb["high_roas_threshold"]:
        pct = rb["high_roas_increase_pct"] / 100
        label = f"ROAS {roas:.2f} > {rb['high_roas_threshold']}"
    elif roas < rb["low_roas_threshold"]:
        if spend > rb["low_roas_high_spend_threshold"]:
            pct = -rb["low_roas_high_spend_decrease_pct"] / 100
            label = f"ROAS {roas:.2f} < {rb['low_roas_threshold']} & spend ${spend:.2f} > ${rb['low_roas_high_spend_threshold']}"
        elif spend >= rb["low_roas_mid_spend_threshold"]:
            pct = -rb["low_roas_mid_spend_decrease_pct"] / 100
            label = f"ROAS {roas:.2f} < {rb['low_roas_threshold']} & spend ${spend:.2f} in [{rb['low_roas_mid_spend_threshold']},{rb['low_roas_high_spend_threshold']}]"
        else:
            pct = -rb["low_roas_low_spend_decrease_pct"] / 100
            label = f"ROAS {roas:.2f} < {rb['low_roas_threshold']} & spend ${spend:.2f} in [{rb['min_spend_for_rules']},{rb['low_roas_mid_spend_threshold']})"
    else:
        pct = 0.0
        label = f"ROAS {roas:.2f} in neutral zone [{rb['low_roas_threshold']},{rb['high_roas_threshold']}]"

    return round(bid * (1 + pct), 2), pct, label


def load_history(client_dir, today):
    """Build per-target change history from past logs/changes_*.csv.

    Same-day logs are excluded: re-running before uploading replaces today's
    batch rather than compounding it. Only action=changed rows count as history —
    held AND rejected rows never actually moved the bid."""
    history = {}
    for path in sorted(glob.glob(os.path.join(client_dir, "logs", "changes_*.csv"))):
        date_part = os.path.basename(path)[len("changes_"):-len(".csv")]
        try:
            log_date = datetime.date.fromisoformat(date_part)
        except ValueError:
            continue
        if log_date >= today:
            continue
        with open(path, newline="") as f:
            for row in csv.DictReader(f):
                if row.get("action", "changed") != "changed":
                    continue
                try:
                    old_bid, new_bid = float(row["old_bid"]), float(row["new_bid"])
                except (ValueError, KeyError):
                    continue
                history.setdefault(row["id"], []).append({
                    "date": log_date,
                    "old_bid": old_bid,
                    "new_bid": new_bid,
                    "direction": "up" if new_bid > old_bid else "down",
                })
    return history


def apply_history_policy(kid, bid, new_bid, tier, history, hist_cfg, today):
    """Checks a proposed change against past batches. Returns (final_bid, hold_reason, note).

    hold_reason is None when the change may proceed; note carries a manual-override
    warning to append to the log either way."""
    past = history.get(kid)
    if not past:
        return new_bid, None, ""

    last = past[-1]
    note = ""
    if abs(last["new_bid"] - bid) > 0.01:
        note = f" [note: engine set {last['new_bid']:.2f} on {last['date']}, file now shows {bid:.2f} — bid was changed outside the engine]"

    if round(new_bid, 2) == round(bid, 2):
        return new_bid, None, note

    days_since = (today - last["date"]).days
    direction = "up" if new_bid > bid else "down"

    if days_since < hist_cfg["cooldown_days"] and tier != "extreme":
        return bid, f"HELD (cooldown): changed {days_since}d ago ({last['old_bid']:.2f} -> {last['new_bid']:.2f}), report window still overlaps that change — waiting {hist_cfg['cooldown_days']}d unless extreme", note

    if direction != last["direction"] and days_since < hist_cfg["reversal_protection_days"] and tier != "extreme":
        return bid, f"HELD (reversal protection): last batch went {last['direction']} on {last['date']}, reversing to {direction} needs {hist_cfg['reversal_protection_days']}d of fresh data or an extreme signal", note

    return new_bid, None, note


def open_campaigns_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = find_campaigns_sheet(wb)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}  # 1-based for openpyxl cell access
    required = ["Entity", "State", "Bid", "Clicks", "Spend", "Sales", "Orders",
                "Operation", "Campaign name (Informational only)", "Ad group name (Informational only)",
                "Keyword text", "Product targeting expression", "Keyword ID", "Product Targeting ID"]
    missing = [c for c in required if c not in idx]
    if missing:
        sys.exit(f"Expected columns missing from sheet '{sheet_name}': {missing}")
    return wb, sheet_name, ws, idx


def propose(client, client_dir):
    """Step 1: compute proposed changes, write a review file. No bulk output yet."""
    cfg = load_config(client_dir)
    input_path = find_input_file(client_dir)
    wb, sheet_name, ws, idx = open_campaigns_sheet(input_path)

    def cell(row_num, col_name):
        return ws.cell(row=row_num, column=idx[col_name]).value

    total_sales, total_orders, total_clicks = 0.0, 0.0, 0.0
    for r in range(2, ws.max_row + 1):
        if cell(r, "Entity") in BID_ENTITIES and cell(r, "State") == "enabled":
            total_clicks += cell(r, "Clicks") or 0
            o = cell(r, "Orders") or 0
            if o > 0:
                total_orders += o
                total_sales += cell(r, "Sales") or 0
    account_aov = (total_sales / total_orders) if total_orders else None
    baseline_cvr = (total_orders / total_clicks) if total_clicks else None

    today = datetime.date.today()
    hist_cfg = cfg.get("history_rules", {"cooldown_days": 7, "reversal_protection_days": 14})
    history = load_history(client_dir, today)

    changes = []
    held = []
    for r in range(2, ws.max_row + 1):
        entity = cell(r, "Entity")
        if entity not in BID_ENTITIES or cell(r, "State") != "enabled":
            continue
        bid = cell(r, "Bid")
        if bid is None:
            continue
        row = {
            "Bid": bid,
            "Clicks": cell(r, "Clicks"),
            "Spend": cell(r, "Spend"),
            "Sales": cell(r, "Sales"),
            "Orders": cell(r, "Orders"),
        }
        kid = cell(r, "Keyword ID") or cell(r, "Product Targeting ID")
        acos_bid, acos_reason, tier = decide_bid(row, cfg, account_aov, baseline_cvr)

        baseline_bid_log = ""
        escalated_log = ""
        if tier is None:
            new_bid, reason = acos_bid, acos_reason
        else:
            base_bid, base_pct, base_label = roas_baseline_bid(bid, row["Spend"] or 0, row["Sales"] or 0, cfg)
            baseline_bid_log = base_bid
            acos_pct = (acos_bid / bid - 1) if bid else 0
            if tier == "extreme":
                new_bid = acos_bid
                escalated_log = "yes"
                reason = f"{acos_reason} [baseline ROAS ({base_label}) would be {base_pct:+.0%} — escalated to ACOS analysis, extreme tier]"
            else:
                new_bid = base_bid
                escalated_log = "no"
                if base_pct == 0:
                    reason = f"Baseline ROAS neutral ({base_label}): no change [ACOS analysis, normal tier, would suggest {acos_pct:+.0%} — {acos_reason}]"
                elif (base_pct > 0) != (acos_pct > 0):
                    reason = f"Baseline ROAS applied ({base_label}): {base_pct:+.0%} [diverges from ACOS analysis, normal tier, opposite direction {acos_pct:+.0%} — {acos_reason}]"
                else:
                    reason = f"Baseline ROAS applied ({base_label}): {base_pct:+.0%} [ACOS analysis, normal tier, would go further: {acos_pct:+.0%} — {acos_reason}]"

        new_bid, hold_reason, note = apply_history_policy(kid, bid, new_bid, tier, history, hist_cfg, today)

        if hold_reason is None and round(new_bid, 2) == round(bid, 2):
            continue
        target_text = cell(r, "Keyword text") or cell(r, "Product targeting expression")
        entry = {
            "row": r,
            "action": "held" if hold_reason else "changed",
            "entity": entity,
            "id": kid,
            "campaign": cell(r, "Campaign name (Informational only)"),
            "ad_group": cell(r, "Ad group name (Informational only)"),
            "target": target_text,
            "old_bid": bid,
            "new_bid": new_bid,
            "baseline_bid": baseline_bid_log,
            "escalated": escalated_log,
            "clicks": row["Clicks"],
            "spend": row["Spend"],
            "sales": row["Sales"],
            "orders": row["Orders"],
            "reason": (hold_reason or reason) + note,
        }
        (held if hold_reason else changes).append(entry)

    date_str = today.isoformat()
    review_dir = os.path.join(client_dir, "review")
    os.makedirs(review_dir, exist_ok=True)
    review_csv = os.path.join(review_dir, f"review_{date_str}.csv")
    review_json = os.path.join(review_dir, f"review_{date_str}.json")

    with open(review_csv, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["approve"] + LOG_FIELDS)
        writer.writeheader()
        for c in changes:
            writer.writerow({"approve": "TRUE", **{k: c[k] for k in LOG_FIELDS}})

    with open(review_json, "w") as f:
        json.dump({
            "source_file": input_path,
            "sheet_name": sheet_name,
            "account_aov": account_aov,
            "baseline_cvr": baseline_cvr,
            "changes": changes,
            "held": held,
        }, f, indent=1, default=str)

    print(f"Account AOV (proxy): {account_aov}")
    print(f"Account baseline CVR: {baseline_cvr:.2%}" if baseline_cvr else "Account baseline CVR: n/a")
    print(f"History: {len(history)} targets with past changes (from {len(set(h['date'] for hs in history.values() for h in hs))} prior batch(es))")
    print(f"Proposed changes: {len(changes)} (edit the 'approve' column, or use the review UI)")
    print(f"Held by history rules: {len(held)} (never need approval — already blocked)")
    print(f"Review file: {review_csv}")
    print(f"Next step: python3 optimize.py {client} --apply {date_str}")


def apply_batch(client, client_dir, date_str):
    """Step 2: apply only the approved rows from a review file, write the final bulk output + log."""
    review_dir = os.path.join(client_dir, "review")
    review_csv = os.path.join(review_dir, f"review_{date_str}.csv")
    review_json = os.path.join(review_dir, f"review_{date_str}.json")
    if not os.path.exists(review_json):
        sys.exit(f"No review found for {date_str} — expected {review_json}")

    with open(review_json) as f:
        state = json.load(f)
    approvals = {}
    with open(review_csv, newline="") as f:
        for row in csv.DictReader(f):
            approvals[row["id"]] = row["approve"].strip().upper() in ("TRUE", "1", "YES")

    source_file = state["source_file"]
    if not os.path.exists(source_file):
        sys.exit(f"Source file used for this review no longer exists: {source_file} — cannot apply.")

    wb, sheet_name, ws, idx = open_campaigns_sheet(source_file)

    final_changed, final_rejected = [], []
    for c in state["changes"]:
        approved = approvals.get(str(c["id"]), True)
        if approved:
            ws.cell(row=c["row"], column=idx["Operation"], value="Update")
            ws.cell(row=c["row"], column=idx["Bid"], value=c["new_bid"])
            final_changed.append(c)
        else:
            rejected = dict(c)
            rejected["action"] = "rejected"
            rejected["reason"] = "Rejected in human review. Original reason: " + rejected["reason"]
            final_rejected.append(rejected)

    out_dir = os.path.join(client_dir, "output")
    log_dir = os.path.join(client_dir, "logs")
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(log_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"bulk_upload_ready_{date_str}.xlsx")
    log_path = os.path.join(log_dir, f"changes_{date_str}.csv")

    wb.save(out_path)

    with open(log_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=LOG_FIELDS)
        writer.writeheader()
        for c in final_changed + final_rejected + state["held"]:
            writer.writerow({k: c[k] for k in LOG_FIELDS})

    increased = sum(1 for c in final_changed if c["new_bid"] > c["old_bid"])
    decreased = sum(1 for c in final_changed if c["new_bid"] < c["old_bid"])
    escalated = sum(1 for c in final_changed if c["escalated"] == "yes")
    print(f"Applied: {len(final_changed)} ({increased} increased, {decreased} decreased), of which {escalated} escalated beyond ROAS baseline (extreme tier)")
    print(f"Rejected in review: {len(final_rejected)}")
    print(f"Held by history rules: {len(state['held'])}")
    print(f"Output: {out_path}")
    print(f"Log: {log_path}")


def main():
    args = sys.argv[1:]
    if len(args) == 1:
        client, apply_date = args[0], None
    elif len(args) == 3 and args[1] == "--apply":
        client, apply_date = args[0], args[2]
    else:
        sys.exit("Usage: python3 optimize.py <client-folder-name> [--apply <YYYY-MM-DD>]")

    base = os.path.dirname(os.path.abspath(__file__))
    client_dir = os.path.join(base, "clients", client)
    if not os.path.isdir(client_dir):
        sys.exit(f"No such client folder: {client_dir}")

    if apply_date is None:
        propose(client, client_dir)
    else:
        apply_batch(client, client_dir, apply_date)


if __name__ == "__main__":
    main()
